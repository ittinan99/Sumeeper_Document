# -*- coding: utf-8 -*-
"""Sumeeper combat sim L1 — deterministic, data from real sheets.
Damage model per GDD "Damage Calculate": dmg = ATK - Armor, floored at 1;
DEF is a depleting pool that eats damage 1:1 and spills the remainder into HP;
DEF refills from base+equipment each combat. Thorn ignores Armor and the floor
but still eats the attacker's DEF pool first.
Rotation is per-action: one Speed Gauge fill = one action by the active
sequence entry, then the rotation advances. Fill rate = SPD of the active
entry. Per-action ATK/SPD/Charge = owner base + piece stat (ADDITIVE, no
substitution); DEF is the only stat summed across the whole loadout.
Player base (starter character, per design): HP10 ATK1 DEF0 SPD1 Charge1 MaxAG2.
On-Exposed fires only when DEF transitions >0 -> 0 (never if it starts at 0).
Action Gauge overflow is discarded on reset.
Assumptions (declared): speed gauge max 10; player wins ties;
feast per-action SPD = Monsters Config base SPD + Feast Sequence entry SPD
(same additive rule as the player side); feast per-action ATK/Charge are the
sequence entry values as-is (Config ATK is legacy, not added); feast DEF pool
= Config DEF + sum of entry DEF; feast HP/MaxAG come from the Feast tab's
hand-set value cells; ability Gain SPD adds to the owner's fill rate for all
actions;
player Special = burst hit with sum of per-action loadout ATK, one attack
(Armor once, floor 1, eats DEF); empty loadout = pure base-stat attack;
On-Half/On-Exposed fire once; Convert/Eater/Scaling logged not simulated;
'Shield' target (none in current data) is treated as Armor; Fury not in any data yet.
"""
import sys
import openpyxl

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")  # monster names are non-ASCII; Windows consoles default to cp874/cp1252

GAUGE_MAX = 10
PLAYER = dict(hp=10, atk=1, dfs=0, spd=1, chg=1, maxag=2)
TICK_LIMIT = 400

def _load(path):
    """Load a workbook with human-friendly errors (for the balance team)."""
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except FileNotFoundError:
        sys.exit(f"[ERROR] หาไฟล์ไม่เจอ: {path}\n  ตรวจว่าโฟลเดอร์ game_document อยู่ครบ")
    except PermissionError:
        sys.exit(f"[ERROR] เปิดไฟล์ไม่ได้: {path}\n  ถ้าไฟล์เปิดค้างอยู่ใน Excel ให้ปิดก่อนแล้วรันใหม่")

# ---------- load equipment ----------
ew = _load(r"D:\Sumeeper\game_document\Sumeeper_Equipment_Sheet.xlsx")
eq = ew["Equipment"]
weapons = {}
for r in range(2, eq.max_row + 1):
    w = eq.cell(r, 2).value
    if not w:
        continue
    weapons[w] = dict(name=w, rar=eq.cell(r, 3).value,
                      atk=eq.cell(r, 5).value or 0, dfs=eq.cell(r, 6).value or 0,
                      spd=eq.cell(r, 7).value or 0, chg=eq.cell(r, 8).value or 0, abil=[])
abs_ = ew["Ability"]
for r in range(2, abs_.max_row + 1):
    w = abs_.cell(r, 1).value
    if not w or w not in weapons:
        continue
    weapons[w]["abil"].append(dict(trig=abs_.cell(r, 2).value, verb=abs_.cell(r, 3).value,
                                   tgt=abs_.cell(r, 4).value, mag=abs_.cell(r, 5).value,
                                   lane=abs_.cell(r, 8).value))

# ---------- load monsters ----------
# Bases come from "Monsters Config" (values), per-action stats from "Feast Sequence";
# only HP and Max AG are read from the "Feast" rollup tab (hand-set value cells, row-aligned
# with Config). The Feast DEF/SPD formula cells (base + SUM of all entries) are the OLD
# entity-sum model and are deliberately NOT used: per-action SPD = base SPD + entry SPD,
# per-action ATK/Charge = entry value as-is (Config ATK is legacy and unused).
mw = _load(r"D:\Sumeeper\game_document\Sumeeper_Monster_Sheet.xlsx")
mc, ft, sq, ca = mw["Monsters Config"], mw["Feast"], mw["Feast Sequence"], mw["Feast Combat Ability"]
monsters = {}
for r in range(2, mc.max_row + 1):
    k = mc.cell(r, 1).value
    if not k:
        continue
    monsters[k] = dict(key=k, name=mc.cell(r, 4).value, tier=int(mc.cell(r, 2).value),
                       hp=ft.cell(r, 5).value, maxag=ft.cell(r, 8).value,
                       dfs=mc.cell(r, 7).value or 0, base_spd=mc.cell(r, 8).value or 0,
                       seq=[], abil=[])
for r in range(3, sq.max_row + 1):
    k = sq.cell(r, 1).value
    if not k:
        continue
    entry = dict(atk=sq.cell(r, 5).value or 0, dfs=sq.cell(r, 6).value or 0,
                 spd=sq.cell(r, 7).value or 0, chg=sq.cell(r, 8).value or 1, abil=[])
    if sq.cell(r, 9).value:
        entry["abil"].append(dict(trig=sq.cell(r, 9).value, verb=sq.cell(r, 10).value,
                                  tgt=sq.cell(r, 11).value, mag=sq.cell(r, 12).value,
                                  lane=sq.cell(r, 14).value))
    monsters[k]["seq"].append(entry)
for m in monsters.values():
    m["dfs"] += sum(e["dfs"] for e in m["seq"])  # DEF pool = base + per-entry contributions
    if m["hp"] is None or m["maxag"] is None:
        sys.exit(f"[ERROR] ค่า HP/Max AG ของ {m['name']} ในแท็บ Feast ว่างอยู่ (แถว {m['key']})\n"
                 f"  มักเกิดเพราะเซลล์เป็นสูตรที่ยังไม่ถูกคำนวณ — เปิดไฟล์ Monster Sheet ใน Excel กด Save หนึ่งครั้งแล้วรันใหม่\n"
                 f"  หรือพิมพ์ค่าเป็นตัวเลขตรงๆ ลงเซลล์นั้นแทนสูตร")
    if not m["seq"]:
        sys.exit(f"[ERROR] {m['name']} ไม่มี action ในแท็บ Feast Sequence เลย — เพิ่มอย่างน้อย 1 แถว")
for r in range(2, ca.max_row + 1):
    k = ca.cell(r, 1).value
    if not k:
        continue
    monsters[k]["abil"].append(dict(trig=ca.cell(r, 3).value, verb=ca.cell(r, 4).value,
                                    tgt=ca.cell(r, 5).value, mag=ca.cell(r, 6).value,
                                    lane=ca.cell(r, 9).value))

# ---------- entity ----------
class Ent:
    def __init__(s, name, hp, dfs, maxag):
        s.name, s.hp, s.maxhp = name, hp, hp
        s.dfs, s.maxag = dfs, maxag
        s.spd = 0  # SPD bonus from abilities; fill rate = active entry spd + this
        s.atk_bonus = 0
        s.thorn = s.armor = 0
        s.sgauge = s.agauge = 0.0
        s.seq_idx = 0
        s.fired = set()
        s.log = []

def apply_ab(a, owner, foe, when, sim):
    if a["trig"] != when:
        return
    verb, tgt, mag = a["verb"], a["tgt"], a["mag"]
    if not isinstance(mag, (int, float)):
        sim.notes.add(f"{owner.name}: Scaling not simulated"); return
    mag = float(mag)
    if verb == "Gain":
        if tgt == "ATK": owner.atk_bonus += mag
        elif tgt == "DEF": owner.dfs += mag
        elif tgt == "SPD": owner.spd += mag
        elif tgt == "Thorn": owner.thorn += mag
        elif tgt in ("Armor", "Shield"): owner.armor += mag  # Shield is legacy wording for Armor
        else: sim.notes.add(f"Gain {tgt} not simulated")
    elif verb == "Restore":
        owner.hp = min(owner.maxhp, owner.hp + mag)
    elif verb == "Lose":
        if tgt == "ATK": owner.atk_bonus -= mag
        elif tgt == "DEF": owner.dfs -= mag
    elif verb in ("Convert", "Eater"):
        sim.notes.add(f"{verb} not simulated ({owner.name})")

def fire(ent, foe, when, sim, once=False, abil=None):
    pool = abil if abil is not None else getattr(ent, "abilities", [])
    for a in pool:
        key = (id(a), when)
        if once and key in ent.fired:
            continue
        if a["trig"] == when:
            if once: ent.fired.add(key)
            apply_ab(a, ent, foe, when, sim)

class Sim:
    def __init__(s):
        s.notes = set()
        s.events = []  # (tick, side) for each Special fired; side = "P"/"M"

def deal(defender, attacker, dmg, sim):
    """DEF pool absorbs 1:1, remainder spills into HP. On-Exposed on >0 -> 0 transition only."""
    if dmg <= 0:
        return 0.0
    had_def = defender.dfs > 0
    absorbed = min(max(defender.dfs, 0.0), dmg)
    defender.dfs -= absorbed
    defender.hp -= dmg - absorbed
    if had_def and defender.dfs <= 0:
        fire(defender, attacker, "On-Exposed", sim, once=True, abil=defender.abilities)
    return dmg

def hit(attacker, defender, atk, sim, t):
    dmg = max(1.0, atk - defender.armor)  # Armor is the only reducer; floor 1
    was_above_half = defender.hp > defender.maxhp / 2
    deal(defender, attacker, dmg, sim)
    # triggers
    fire(attacker, defender, "On-Hit", sim, abil=attacker.cur_abil)
    fire(defender, attacker, "On-Damaged", sim, abil=defender.abilities)
    if was_above_half and defender.hp <= defender.maxhp / 2:
        fire(defender, attacker, "On-Half", sim, once=True, abil=defender.abilities)
    if defender.thorn > 0:
        deal(attacker, defender, defender.thorn, sim)  # no Armor cut, no floor
    return dmg

def step(ent, foe, sim, t):
    """One gauge fill = one action by the active sequence entry, then rotate."""
    entry = ent.sequence[ent.seq_idx]
    ent.seq_idx = (ent.seq_idx + 1) % len(ent.sequence)
    ent.cur_abil = ent.abilities + entry.get("abil", [])
    hit(ent, foe, entry["atk"] + ent.atk_bonus, sim, t)
    ent.agauge += entry["chg"]
    if ent.agauge >= ent.maxag and foe.hp > 0 and ent.hp > 0:
        ent.agauge = 0
        do_special(ent, foe, sim, t)
    ent.cur_abil = ent.abilities

def do_special(ent, foe, sim, t):
    sim.events.append((t, "P" if ent.name == "player" else "M"))
    sp = [a for a in ent.abilities if a.get("lane") == "Special"]
    if sp:
        for a in sp: apply_ab(a, ent, foe, "Special", sim)
    else:  # player default burst
        ent.cur_abil = []
        dmg = max(1.0, ent.burst_atk + ent.atk_bonus - foe.armor)
        deal(foe, ent, dmg, sim)
    ent.specials += 1

def fight(loadout_names, mkey):
    sim = Sim()
    m = monsters[mkey]
    pw = [weapons[n] for n in loadout_names]
    p = Ent("player", PLAYER["hp"], PLAYER["dfs"] + sum(w["dfs"] for w in pw), PLAYER["maxag"])
    # per-action values = base + piece stat (additive, per GDD); empty loadout -> pure base attack
    if pw:
        p.sequence = [dict(atk=PLAYER["atk"] + w["atk"], spd=PLAYER["spd"] + w["spd"],
                           chg=PLAYER["chg"] + w["chg"],
                           abil=[a for a in w["abil"] if a["lane"] == "Action"]) for w in pw]
    else:
        p.sequence = [dict(atk=PLAYER["atk"], spd=PLAYER["spd"], chg=PLAYER["chg"], abil=[])]
    p.abilities = [a for w in pw for a in w["abil"] if a["lane"] != "Action"]
    p.burst_atk = sum(e["atk"] for e in p.sequence)
    e = Ent(m["name"], m["hp"], m["dfs"], m["maxag"])
    e.sequence = [dict(x, spd=m["base_spd"] + x["spd"]) for x in m["seq"]]  # per-action SPD = base + entry (same rule as player)
    e.abilities = m["abil"]
    e.burst_atk = 0
    p.specials = e.specials = 0
    p.cur_abil = p.abilities; e.cur_abil = e.abilities
    fire(p, e, "On-Start", sim, once=True, abil=p.abilities)
    fire(e, p, "On-Start", sim, once=True, abil=e.abilities)
    t = 0
    history = [(0, p.hp, p.dfs, e.hp, e.dfs)]  # (tick, pHP, pDEF, mHP, mDEF)
    while t < TICK_LIMIT and p.hp > 0 and e.hp > 0:
        t += 1
        # fill rate = SPD of the active (next-to-act) entry + any SPD bonus; floored at 1 so rotation never stalls
        p.sgauge += max(1, p.sequence[p.seq_idx]["spd"] + p.spd)
        e.sgauge += max(1, e.sequence[e.seq_idx]["spd"] + e.spd)
        if p.sgauge >= GAUGE_MAX:
            p.sgauge -= GAUGE_MAX
            step(p, e, sim, t)
        if e.hp <= 0 or p.hp <= 0:
            history.append((t, p.hp, p.dfs, e.hp, e.dfs))
            break
        if e.sgauge >= GAUGE_MAX:
            e.sgauge -= GAUGE_MAX
            step(e, p, sim, t)
        history.append((t, p.hp, p.dfs, e.hp, e.dfs))
    win = "P" if e.hp <= 0 else ("M" if p.hp <= 0 else "TIMEOUT")
    return dict(win=win, ticks=t, php=round(p.hp, 1), mhp=round(e.hp, 1),
                pmaxhp=p.maxhp, mmaxhp=e.maxhp,
                psp=p.specials, msp=e.specials, notes=sim.notes,
                history=history, events=sim.events)

# ---------- archetype loadouts per monster-tier gear stage ----------
BUILDS = {
    1: {"Heavy-1": ["Spirit Fork"],
        "Rush-3":  ["Mysterious Pot", "Broken Plate", "Mortar and Pestle"],
        "Bal-2":   ["Frying pan", "Chopping Board"]},
    2: {"Heavy-1": ["Cast Iron Lid"],
        "Rush-3":  ["Mysterious Pot", "Cloud Cutting Spoon", "Windmill Spatula"],
        "Bal-2":   ["Spirit Fork", "Chopping Board"]},
    3: {"Heavy-1": ["Darkness Apron"],
        "Rush-3":  ["Bottomless Chef's Hat", "Cloud Cutting Spoon", "Infinity Gloves"],
        "Bal-2":   ["Infinity Gloves", "Chopping Board"]},
    4: {"Heavy-1": ["Butcher Cleaver"],
        "Rush-3":  ["Bottomless Chef's Hat", "Miracle Knife", "Infinity Gloves"],
        "Bal-2":   ["Emperor Knife", "Immortal Ladle"]},
    5: {"Heavy-1": ["Titan Stone Mortar"],
        "Rush-3":  ["Gaint Fork", "Miracle Knife", "Maid Apron"],
        "Bal-2":   ["Emperor Knife", "Dragon Bone Ladle"]},
}

order = ["MONSTER_ORANGECAT","MONSTER_ONPIRENION","MONSTER_MUSHROOMRAT","MONSTER_CABBAGEDOG",
         "MONSTER_CHERRYBIRD","MONSTER_PUDDINGCRAP","MONSTER_SUGARTIGER",
         "MONSTER_SPICYHEDGEHOG","MONSTER_BLUEFIN","MONSTER_PIZZATURTLE","MONSTER_DURAINBOMB","MONSTER_CHIVE",
         "MONSTER_BROCOLION","MONSTER_BANANAOCTOPUS","MONSTER_UBEECORN","MONSTER_MONKEYCUPSSPICES",
         "MONSTER_DEARBOKCHOY","MONSTER_JELLYUDON"]

if __name__ == "__main__":
    allnotes = set()
    print(f"{'Monster':18}{'T':>2} | " + " | ".join(f"{b:^24}" for b in ["Heavy-1","Rush-3","Bal-2"]))
    print("-" * 100)
    for mk in order:
        m = monsters[mk]
        cells = []
        for bname in ["Heavy-1", "Rush-3", "Bal-2"]:
            res = fight(BUILDS[m["tier"]][bname], mk)
            allnotes |= res["notes"]
            tag = res["win"]
            cells.append(f"{tag} t{res['ticks']:<3} pHP{res['php']:>5} mHP{res['mhp']:>5}")
        print(f"{m['name'][:18]:18}{m['tier']:>2} | " + " | ".join(f"{c:24}" for c in cells))
    print()
    print("NOT SIMULATED:", sorted(allnotes))


